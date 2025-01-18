import openpyxl
import json
import csv

def read_json(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

def read_csv(file_path):
    with open(file_path, newline='') as file:
        return list(csv.DictReader(file))

def read_excel(file_path, test_sheet):
    spreadSheet = openpyxl.load_workbook(file_path)
    sheetData = spreadSheet[test_sheet]
    maxRow = sheetData.max_row
    maxCol = sheetData.max_column
    excelData = []
    for i in range(2, maxRow + 1):
        row = ()
        for j in range(1, maxCol + 1):
            cellData = sheetData.cell(row = i, column = j)
            row += (cellData.value,)
        excelData.append(row)
    spreadSheet.close()
    return excelData